from datetime import date, timedelta
from openpyxl import load_workbook
from tkinter import filedialog
import openpyxl
import datetime
import calendar
import zipfile
import shutil
import time
import os
import sys


#
print("""  		    _       _____   ____   _   _   _____    ____   	  _____  __  __
            /\     | |     |  ___| / __ \ | \ | | / ____|  / __ \ 	 / ____||  \/  |
           /  \    | |     | |_	  | |  | ||  \| || (___   | |  | |	| |  __ | |\/| |
          / /\ \   | |     |  _|  | |  | || . \`| \___ \  | |  | |	| | |_ || |  | |
         / ____ \  | |___  | |	  | |__| || |\  | ____) | | |__| |	| |__| || |  | |
        /_/    \_\ |_____| |_|     \____/ |_| \_||_____/   \____/ 	 \_____||_|  |_|

        ----------------------------------------------------------------------------
        
        """)





print("\n" + ".................................................." + "\n")
print("// Código creado por Alfonso Gómez Medel @ 2026 //"+"\n")



# 
def limpiar_caracteres(texto):
    caracteres_no_validos = '\x00\x01\x02\x03\x04\x05\x06\x07\x08\x0B\x0C\x0E\x0F\x10\x11\x12\x13\x14\x15\x16\x17\x18\x19\x1A\x1B\x1C\x1D\x1E\x1F'
    return ''.join(char for char in texto if char not in caracteres_no_validos)

    

def dividir_por_anchos(linea, anchos_columnas):
    
    columnas = []
    
    indice_inicio = 0
    for ancho in anchos_columnas:

        columna = linea[indice_inicio:indice_inicio+ancho].strip() 
        columnas.append(columna)

        indice_inicio += ancho
    return columnas



archivo = filedialog.askopenfilename()

anchos_columnas = [4, 5, 30,15, 60, 85, 95, 105]  

anchos_excel = {
    'A': 5, 
    'B': 5,  
    'C': 15,  
    'D': 15,
    'E': 25,
    'F': 5,
    'G': 5,
    'H': 5
}

with open(f'{archivo}', 'r', encoding='latin-1') as file_in:
    lines = file_in.readlines()

workbook = openpyxl.Workbook()
sheet = workbook.active


for row_idx, line in enumerate(lines, start=1):
    cleaned_line = limpiar_caracteres(line.strip())
    
    columnas = dividir_por_anchos(cleaned_line, anchos_columnas)
    
    for col_idx, value in enumerate(columnas, start=1):
        sheet.cell(row=row_idx, column=col_idx, value=value)


for columna, ancho in anchos_excel.items():
    sheet.column_dimensions[columna].width = ancho

output = str(archivo)[:-4]
workbook.save(f'{output}.xlsx')
workbook.close()


print("\n"+"Los ficheros se han separado en columnas con éxito.")
time.sleep(3)




#
datos00 = f"{output}.xlsx"

libro_datos = load_workbook(filename=datos00)

hoja_datos = libro_datos.active



filas_a_eliminar_temp = []

for fila in range(1, hoja_datos.max_row + 1):
    valor_celda = hoja_datos[f"A{fila}"].value
    
    
    if valor_celda is None or str(valor_celda).strip() == "":
        filas_a_eliminar_temp.append(fila)
    
    
    if valor_celda and not (str(valor_celda).startswith("011") or #Nº ASP
                            str(valor_celda).startswith("012") or
                            str(valor_celda).startswith("013") or
                            str(valor_celda).startswith("015")):
        
        filas_a_eliminar_temp.append(fila)




for fila_a_eliminar in reversed(filas_a_eliminar_temp):
    hoja_datos.delete_rows(fila_a_eliminar)
    



nombre_archivo = os.path.basename(output)

mapa_dias = {
    "_L": "Lunes",
    "_M": "Martes",
    "_X": "Miércoles",
    "_J": "Jueves",
    "_V": "Viernes",
    "_S": "Sábado",
    "_D": "Domingo"
}

dia_semana = ""
fecha_usuario = ""


nombre_base = nombre_archivo

for sufijo, dia in mapa_dias.items():
    if nombre_archivo.endswith(sufijo):
        dia_semana = dia
        nombre_base = nombre_archivo[:-len(sufijo)]
        break
    

if dia_semana:
    hoy = date.today()

    mapa_weekday = {
        "Lunes": 0,
        "Martes": 1,
        "Miércoles": 2,
        "Jueves": 3,
        "Viernes": 4,
        "Sábado": 5,
        "Domingo": 6
    }

    target_weekday = mapa_weekday[dia_semana]

    lunes_semana = hoy - timedelta(days=hoy.weekday())

    fecha_candidata = lunes_semana + timedelta(days=target_weekday)

    if fecha_candidata <= hoy:
        fecha_candidata += timedelta(weeks=1)

    fecha_usuario = fecha_candidata.strftime("%d/%m/%Y")






nombre_base = os.path.splitext(nombre_base)[0]

nombre_base = nombre_base.strip().title()

for fila in range(1, hoja_datos.max_row + 1):

    if fecha_usuario:
        hoja_datos.cell(row=fila, column=6, value=fecha_usuario)

    hoja_datos.cell(row=fila, column=7, value=nombre_base)


# os.rename(f"{output}.xlsx",f"{output}_backup.xlsx")
libro_datos.save(filename=f"{output}.xlsx")
libro_datos.close()

print("\n"+"Listo!")
exit(0)